import pandas as pd
import streamlit as st
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

def day_fraction(date):
    if pd.isnull(date):
        return 0
    days_in_month = (date.replace(month=(date.month % 12 + 1), day=1) - pd.Timedelta(days=1)).day
    return date.day / days_in_month

def detect_columns(df):
    def find_col(possible_keywords):
        for col in df.columns:
            col_str = str(col).lower().strip()
            if any(kw in col_str for kw in possible_keywords):
                return col
        return None
    planned_start = find_col(['planned start', 'start date', 'planned begin'])
    planned_end = find_col(['planned end', 'end date', 'planned finish'])
    actual_start = find_col(['actual start', 'start actual'])
    actual_end = find_col(['actual end', 'actual finish'])
    return planned_start, planned_end, actual_start, actual_end

def detect_subtasks(df):
    if 'Task S. No' in df.columns:
        return df['Task S. No'].astype(str).str.contains(r'^\d+\.\d+$')
    else:
        return pd.Series([False] * len(df))

def draw_colored_cells(ws, df, months, base_col, row_offset, ps_col, pe_col, as_col, ae_col):
    for row in range(len(df)):
        planned_start = df[ps_col].iloc[row] if ps_col else None
        planned_end = df[pe_col].iloc[row] if pe_col else None
        actual_start = df[as_col].iloc[row] if as_col else None
        actual_end = df[ae_col].iloc[row] if ae_col else None
        is_sub = df['is_subtask'].iloc[row]

        planned_color = "90EE90" if is_sub else "00CC00"
        actual_color = "800080" if is_sub else "0000FF"

        for idx, month in enumerate(months):
            cell_col_letter = get_column_letter(base_col + idx)
            cell = ws[f"{cell_col_letter}{row_offset + row}"]
            fill_color = None

            def in_range(start, end):
                if pd.isnull(start) or pd.isnull(end):
                    return False
                return start.replace(day=1) <= month <= end.replace(day=1)

            if in_range(planned_start, planned_end):
                fill_color = planned_color
            if in_range(actual_start, actual_end):
                fill_color = actual_color

            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

def process_excel(file):
    xls = pd.ExcelFile(file)
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            ps_col, pe_col, as_col, ae_col = detect_columns(df)

            if not any([ps_col, pe_col, as_col, ae_col]):
                continue

            for col in [ps_col, pe_col, as_col, ae_col]:
                if col:
                    df[col] = pd.to_datetime(df[col], errors='coerce')

            df['is_subtask'] = detect_subtasks(df)

            min_date = min([df[col].min(skipna=True) for col in [ps_col, as_col] if col], default=pd.NaT)
            max_date = max([df[col].max(skipna=True) for col in [pe_col, ae_col] if col], default=pd.NaT)

            if pd.isnull(min_date) or pd.isnull(max_date):
                continue

            months = pd.date_range(start=min_date.replace(day=1), end=max_date.replace(day=1), freq='MS')
            base_col = len(df.columns) + 2
            row_offset = 2

            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

            ws = writer.sheets[sheet_name]
            for idx, month in enumerate(months):
                col_letter = get_column_letter(base_col + idx)
                ws[f"{col_letter}1"] = month.strftime('%b %Y')
                ws.column_dimensions[col_letter].width = 10

            draw_colored_cells(ws, df, months, base_col, row_offset, ps_col, pe_col, as_col, ae_col)
        except Exception as e:
            print(f"Error in sheet {sheet_name}: {e}")
    writer.close()
    output.seek(0)
    return output

# ---------------- Streamlit UI ----------------
st.title("📊 Excel Gantt Chart Generator")
st.write("Upload an Excel file with Planned/Actual Start and End dates to visualize Gantt charts.")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file:
    with st.spinner("Processing..."):
        result_file = process_excel(uploaded_file)
        st.success("✅ Gantt chart Excel generated!")
        st.download_button(
            label="📥 Download Gantt Excel",
            data=result_file,
            file_name="gantt_chart_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
